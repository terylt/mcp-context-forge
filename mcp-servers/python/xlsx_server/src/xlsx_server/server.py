#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Location: ./mcp-servers/python/xlsx_server/src/xlsx_server/server.py
Copyright 2025
SPDX-License-Identifier: Apache-2.0
Authors: Mihai Criveti

XLSX MCP Server

A comprehensive MCP server for creating, editing, and analyzing Microsoft Excel (.xlsx) spreadsheets.
Provides tools for workbook creation, data manipulation, formatting, formulas, and spreadsheet analysis.
"""

import asyncio
import json
import logging
import sys
from pathlib import Path
from typing import Any, Sequence

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from mcp.server import Server
from mcp.server.models import InitializationOptions
from mcp.types import EmbeddedResource, ImageContent, TextContent, Tool
from pydantic import BaseModel, Field

# Configure logging to stderr to avoid MCP protocol interference
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[logging.StreamHandler(sys.stderr)],
)
logger = logging.getLogger(__name__)

# Create server instance
server = Server("xlsx-server")


class WorkbookRequest(BaseModel):
    """Base request for workbook operations."""
    file_path: str = Field(..., description="Path to the XLSX file")


class CreateWorkbookRequest(WorkbookRequest):
    """Request to create a new workbook."""
    sheet_names: list[str] | None = Field(None, description="Names of sheets to create")


class WriteDataRequest(WorkbookRequest):
    """Request to write data to a worksheet."""
    sheet_name: str | None = Field(None, description="Sheet name (uses active sheet if None)")
    data: list[list[Any]] = Field(..., description="Data to write (2D array)")
    start_row: int = Field(1, description="Starting row (1-indexed)")
    start_col: int = Field(1, description="Starting column (1-indexed)")
    headers: list[str] | None = Field(None, description="Column headers")


class ReadDataRequest(WorkbookRequest):
    """Request to read data from a worksheet."""
    sheet_name: str | None = Field(None, description="Sheet name (uses active sheet if None)")
    start_row: int | None = Field(None, description="Starting row to read")
    end_row: int | None = Field(None, description="Ending row to read")
    start_col: int | None = Field(None, description="Starting column to read")
    end_col: int | None = Field(None, description="Ending column to read")


class FormatCellsRequest(WorkbookRequest):
    """Request to format cells."""
    sheet_name: str | None = Field(None, description="Sheet name")
    cell_range: str = Field(..., description="Cell range (e.g., 'A1:C5')")
    font_name: str | None = Field(None, description="Font name")
    font_size: int | None = Field(None, description="Font size")
    font_bold: bool | None = Field(None, description="Bold font")
    font_italic: bool | None = Field(None, description="Italic font")
    font_color: str | None = Field(None, description="Font color (hex)")
    background_color: str | None = Field(None, description="Background color (hex)")
    alignment: str | None = Field(None, description="Text alignment")


class AddFormulaRequest(WorkbookRequest):
    """Request to add a formula to a cell."""
    sheet_name: str | None = Field(None, description="Sheet name")
    cell: str = Field(..., description="Cell reference (e.g., 'A1')")
    formula: str = Field(..., description="Formula to add")


class AnalyzeWorkbookRequest(WorkbookRequest):
    """Request to analyze workbook content."""
    include_structure: bool = Field(True, description="Include workbook structure analysis")
    include_data_summary: bool = Field(True, description="Include data summary")
    include_formulas: bool = Field(True, description="Include formula analysis")


class SpreadsheetOperation:
    """Handles spreadsheet operations."""

    @staticmethod
    def create_workbook(file_path: str, sheet_names: list[str] | None = None) -> dict[str, Any]:
        """Create a new XLSX workbook."""
        try:
            # Create workbook
            wb = Workbook()

            # Remove default sheet if we're creating custom ones
            if sheet_names:
                # Remove default sheet
                wb.remove(wb.active)

                # Create named sheets
                for sheet_name in sheet_names:
                    wb.create_sheet(title=sheet_name)
            else:
                # Rename default sheet
                wb.active.title = "Sheet1"

            # Ensure directory exists
            Path(file_path).parent.mkdir(parents=True, exist_ok=True)

            # Save workbook
            wb.save(file_path)

            return {
                "success": True,
                "message": f"Workbook created at {file_path}",
                "file_path": file_path,
                "sheets": [sheet.title for sheet in wb.worksheets],
                "total_sheets": len(wb.worksheets)
            }
        except Exception as e:
            logger.error(f"Error creating workbook: {e}")
            return {"success": False, "error": str(e)}

    @staticmethod
    def write_data(file_path: str, data: list[list[Any]], sheet_name: str | None = None,
                   start_row: int = 1, start_col: int = 1, headers: list[str] | None = None) -> dict[str, Any]:
        """Write data to a worksheet."""
        try:
            if not Path(file_path).exists():
                return {"success": False, "error": f"Workbook not found: {file_path}"}

            wb = openpyxl.load_workbook(file_path)

            # Get worksheet
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    ws = wb.create_sheet(title=sheet_name)
                else:
                    ws = wb[sheet_name]
            else:
                ws = wb.active

            # Write headers if provided
            current_row = start_row
            if headers:
                for col_idx, header in enumerate(headers):
                    ws.cell(row=current_row, column=start_col + col_idx, value=header)
                    # Make headers bold
                    ws.cell(row=current_row, column=start_col + col_idx).font = Font(bold=True)
                current_row += 1

            # Write data
            for row_idx, row_data in enumerate(data):
                for col_idx, cell_value in enumerate(row_data):
                    ws.cell(row=current_row + row_idx, column=start_col + col_idx, value=cell_value)

            wb.save(file_path)

            return {
                "success": True,
                "message": f"Data written to {sheet_name or 'active sheet'}",
                "sheet_name": ws.title,
                "rows_written": len(data),
                "cols_written": max(len(row) for row in data) if data else 0,
                "start_cell": f"{get_column_letter(start_col)}{start_row}",
                "has_headers": bool(headers)
            }
        except Exception as e:
            logger.error(f"Error writing data: {e}")
            return {"success": False, "error": str(e)}

    @staticmethod
    def read_data(file_path: str, sheet_name: str | None = None, start_row: int | None = None,
                  end_row: int | None = None, start_col: int | None = None, end_col: int | None = None) -> dict[str, Any]:
        """Read data from a worksheet."""
        try:
            if not Path(file_path).exists():
                return {"success": False, "error": f"Workbook not found: {file_path}"}

            wb = openpyxl.load_workbook(file_path)

            # Get worksheet
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
                ws = wb[sheet_name]
            else:
                ws = wb.active

            # Determine data range
            if not start_row:
                start_row = 1
            if not end_row:
                end_row = ws.max_row
            if not start_col:
                start_col = 1
            if not end_col:
                end_col = ws.max_column

            # Read data
            data = []
            for row in ws.iter_rows(min_row=start_row, max_row=end_row,
                                   min_col=start_col, max_col=end_col, values_only=True):
                data.append(list(row))

            return {
                "success": True,
                "sheet_name": ws.title,
                "data": data,
                "rows_read": len(data),
                "cols_read": end_col - start_col + 1,
                "range": f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
            }
        except Exception as e:
            logger.error(f"Error reading data: {e}")
            return {"success": False, "error": str(e)}

    @staticmethod
    def format_cells(file_path: str, cell_range: str, sheet_name: str | None = None,
                     font_name: str | None = None, font_size: int | None = None,
                     font_bold: bool | None = None, font_italic: bool | None = None,
                     font_color: str | None = None, background_color: str | None = None,
                     alignment: str | None = None) -> dict[str, Any]:
        """Format cells in a worksheet."""
        try:
            if not Path(file_path).exists():
                return {"success": False, "error": f"Workbook not found: {file_path}"}

            wb = openpyxl.load_workbook(file_path)

            # Get worksheet
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
                ws = wb[sheet_name]
            else:
                ws = wb.active

            # Apply formatting to range
            cell_range_obj = ws[cell_range]

            # Handle single cell vs range
            if hasattr(cell_range_obj, '__iter__') and not isinstance(cell_range_obj, openpyxl.cell.Cell):
                # Range of cells
                cells = []
                for row in cell_range_obj:
                    if hasattr(row, '__iter__'):
                        cells.extend(row)
                    else:
                        cells.append(row)
            else:
                # Single cell
                cells = [cell_range_obj]

            # Apply formatting
            for cell in cells:
                # Font formatting
                font_kwargs = {}
                if font_name:
                    font_kwargs['name'] = font_name
                if font_size:
                    font_kwargs['size'] = font_size
                if font_bold is not None:
                    font_kwargs['bold'] = font_bold
                if font_italic is not None:
                    font_kwargs['italic'] = font_italic
                if font_color:
                    font_kwargs['color'] = font_color.replace('#', '')

                if font_kwargs:
                    cell.font = Font(**font_kwargs)

                # Background color
                if background_color:
                    cell.fill = PatternFill(start_color=background_color.replace('#', ''),
                                          end_color=background_color.replace('#', ''),
                                          fill_type="solid")

                # Alignment
                if alignment:
                    alignment_map = {
                        'left': 'left', 'center': 'center', 'right': 'right',
                        'top': 'top', 'middle': 'center', 'bottom': 'bottom'
                    }
                    if alignment.lower() in alignment_map:
                        cell.alignment = Alignment(horizontal=alignment_map[alignment.lower()])

            wb.save(file_path)

            return {
                "success": True,
                "message": f"Formatting applied to range {cell_range}",
                "sheet_name": ws.title,
                "cell_range": cell_range,
                "formatting_applied": {
                    "font_name": font_name,
                    "font_size": font_size,
                    "font_bold": font_bold,
                    "font_italic": font_italic,
                    "font_color": font_color,
                    "background_color": background_color,
                    "alignment": alignment
                }
            }
        except Exception as e:
            logger.error(f"Error formatting cells: {e}")
            return {"success": False, "error": str(e)}

    @staticmethod
    def add_formula(file_path: str, cell: str, formula: str, sheet_name: str | None = None) -> dict[str, Any]:
        """Add a formula to a cell."""
        try:
            if not Path(file_path).exists():
                return {"success": False, "error": f"Workbook not found: {file_path}"}

            wb = openpyxl.load_workbook(file_path)

            # Get worksheet
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
                ws = wb[sheet_name]
            else:
                ws = wb.active

            # Add formula
            if not formula.startswith('='):
                formula = '=' + formula

            ws[cell] = formula

            wb.save(file_path)

            return {
                "success": True,
                "message": f"Formula added to cell {cell}",
                "sheet_name": ws.title,
                "cell": cell,
                "formula": formula
            }
        except Exception as e:
            logger.error(f"Error adding formula: {e}")
            return {"success": False, "error": str(e)}

    @staticmethod
    def analyze_workbook(file_path: str, include_structure: bool = True, include_data_summary: bool = True,
                        include_formulas: bool = True) -> dict[str, Any]:
        """Analyze workbook content and structure."""
        try:
            if not Path(file_path).exists():
                return {"success": False, "error": f"Workbook not found: {file_path}"}

            wb = openpyxl.load_workbook(file_path)
            analysis = {"success": True}

            if include_structure:
                structure = {
                    "total_sheets": len(wb.worksheets),
                    "sheet_names": [sheet.title for sheet in wb.worksheets],
                    "active_sheet": wb.active.title,
                    "sheets_info": []
                }

                for sheet in wb.worksheets:
                    sheet_info = {
                        "name": sheet.title,
                        "max_row": sheet.max_row,
                        "max_column": sheet.max_column,
                        "data_range": f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}",
                        "has_data": sheet.max_row > 0 and sheet.max_column > 0
                    }
                    structure["sheets_info"].append(sheet_info)

                analysis["structure"] = structure

            if include_data_summary:
                data_summary = {}

                for sheet in wb.worksheets:
                    sheet_summary = {
                        "total_cells": sheet.max_row * sheet.max_column,
                        "non_empty_cells": 0,
                        "data_types": {"text": 0, "number": 0, "formula": 0, "date": 0, "boolean": 0},
                        "sample_data": []
                    }

                    # Sample first 5 rows of data
                    sample_rows = min(5, sheet.max_row)
                    for row in sheet.iter_rows(min_row=1, max_row=sample_rows, values_only=True):
                        sheet_summary["sample_data"].append(list(row))

                    # Count data types and non-empty cells
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value is not None:
                                sheet_summary["non_empty_cells"] += 1

                                if hasattr(cell, 'data_type'):
                                    if cell.data_type == 'f':
                                        sheet_summary["data_types"]["formula"] += 1
                                    elif cell.data_type == 'n':
                                        sheet_summary["data_types"]["number"] += 1
                                    elif cell.data_type == 'd':
                                        sheet_summary["data_types"]["date"] += 1
                                    elif cell.data_type == 'b':
                                        sheet_summary["data_types"]["boolean"] += 1
                                    else:
                                        sheet_summary["data_types"]["text"] += 1

                    data_summary[sheet.title] = sheet_summary

                analysis["data_summary"] = data_summary

            if include_formulas:
                formulas = {}

                for sheet in wb.worksheets:
                    sheet_formulas = []
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                                sheet_formulas.append({
                                    "cell": cell.coordinate,
                                    "formula": cell.value,
                                    "value": cell.displayed_value if hasattr(cell, 'displayed_value') else None
                                })

                    if sheet_formulas:
                        formulas[sheet.title] = sheet_formulas

                analysis["formulas"] = formulas

            return analysis
        except Exception as e:
            logger.error(f"Error analyzing workbook: {e}")
            return {"success": False, "error": str(e)}

    @staticmethod
    def create_chart(file_path: str, sheet_name: str | None = None, chart_type: str = "column",
                     data_range: str = "", title: str = "", x_axis_title: str = "",
                     y_axis_title: str = "") -> dict[str, Any]:
        """Create a chart in a worksheet."""
        try:
            if not Path(file_path).exists():
                return {"success": False, "error": f"Workbook not found: {file_path}"}

            wb = openpyxl.load_workbook(file_path)

            # Get worksheet
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
                ws = wb[sheet_name]
            else:
                ws = wb.active

            # Import chart classes
            from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart
            from openpyxl.chart.reference import Reference

            # Create chart based on type
            chart_classes = {
                "column": BarChart,
                "bar": BarChart,
                "line": LineChart,
                "pie": PieChart,
                "scatter": ScatterChart
            }

            if chart_type not in chart_classes:
                return {"success": False, "error": f"Unsupported chart type: {chart_type}"}

            chart = chart_classes[chart_type]()

            # Set chart properties
            if title:
                chart.title = title
            if x_axis_title and hasattr(chart, 'x_axis'):
                chart.x_axis.title = x_axis_title
            if y_axis_title and hasattr(chart, 'y_axis'):
                chart.y_axis.title = y_axis_title

            # Add data if range provided
            if data_range:
                data = Reference(ws, range_string=data_range)
                chart.add_data(data, titles_from_data=True)

            # Add chart to worksheet
            ws.add_chart(chart, "E2")  # Default position

            wb.save(file_path)

            return {
                "success": True,
                "message": f"Chart created in {ws.title}",
                "sheet_name": ws.title,
                "chart_type": chart_type,
                "data_range": data_range,
                "title": title
            }
        except Exception as e:
            logger.error(f"Error creating chart: {e}")
            return {"success": False, "error": str(e)}


@server.list_tools()
async def handle_list_tools() -> list[Tool]:
    """List available XLSX tools."""
    return [
        Tool(
            name="create_workbook",
            description="Create a new XLSX workbook",
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Path where the workbook will be saved"
                    },
                    "sheet_names": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Names of sheets to create (optional)"
                    }
                },
                "required": ["file_path"]
            }
        ),
        Tool(
            name="write_data",
            description="Write data to a worksheet",
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Path to the XLSX file"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Sheet name (optional, uses active sheet if not specified)"
                    },
                    "data": {
                        "type": "array",
                        "items": {
                            "type": "array",
                            "items": {}
                        },
                        "description": "Data to write (2D array)"
                    },
                    "start_row": {
                        "type": "integer",
                        "description": "Starting row (1-indexed)",
                        "default": 1
                    },
                    "start_col": {
                        "type": "integer",
                        "description": "Starting column (1-indexed)",
                        "default": 1
                    },
                    "headers": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Column headers (optional)"
                    }
                },
                "required": ["file_path", "data"]
            }
        ),
        Tool(
            name="read_data",
            description="Read data from a worksheet",
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Path to the XLSX file"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Sheet name (optional, uses active sheet if not specified)"
                    },
                    "start_row": {
                        "type": "integer",
                        "description": "Starting row to read (optional)"
                    },
                    "end_row": {
                        "type": "integer",
                        "description": "Ending row to read (optional)"
                    },
                    "start_col": {
                        "type": "integer",
                        "description": "Starting column to read (optional)"
                    },
                    "end_col": {
                        "type": "integer",
                        "description": "Ending column to read (optional)"
                    }
                },
                "required": ["file_path"]
            }
        ),
        Tool(
            name="format_cells",
            description="Format cells in a worksheet",
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Path to the XLSX file"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Sheet name (optional)"
                    },
                    "cell_range": {
                        "type": "string",
                        "description": "Cell range to format (e.g., 'A1:C5')"
                    },
                    "font_name": {
                        "type": "string",
                        "description": "Font name (optional)"
                    },
                    "font_size": {
                        "type": "integer",
                        "description": "Font size (optional)"
                    },
                    "font_bold": {
                        "type": "boolean",
                        "description": "Bold font (optional)"
                    },
                    "font_italic": {
                        "type": "boolean",
                        "description": "Italic font (optional)"
                    },
                    "font_color": {
                        "type": "string",
                        "description": "Font color in hex format (optional)"
                    },
                    "background_color": {
                        "type": "string",
                        "description": "Background color in hex format (optional)"
                    },
                    "alignment": {
                        "type": "string",
                        "description": "Text alignment (left, center, right, top, middle, bottom)"
                    }
                },
                "required": ["file_path", "cell_range"]
            }
        ),
        Tool(
            name="add_formula",
            description="Add a formula to a cell",
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Path to the XLSX file"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Sheet name (optional)"
                    },
                    "cell": {
                        "type": "string",
                        "description": "Cell reference (e.g., 'A1')"
                    },
                    "formula": {
                        "type": "string",
                        "description": "Formula to add (with or without leading =)"
                    }
                },
                "required": ["file_path", "cell", "formula"]
            }
        ),
        Tool(
            name="analyze_workbook",
            description="Analyze workbook content, structure, and formulas",
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Path to the XLSX file"
                    },
                    "include_structure": {
                        "type": "boolean",
                        "description": "Include workbook structure analysis",
                        "default": True
                    },
                    "include_data_summary": {
                        "type": "boolean",
                        "description": "Include data summary",
                        "default": True
                    },
                    "include_formulas": {
                        "type": "boolean",
                        "description": "Include formula analysis",
                        "default": True
                    }
                },
                "required": ["file_path"]
            }
        ),
        Tool(
            name="create_chart",
            description="Create a chart in a worksheet",
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Path to the XLSX file"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Sheet name (optional)"
                    },
                    "chart_type": {
                        "type": "string",
                        "enum": ["column", "bar", "line", "pie", "scatter"],
                        "description": "Type of chart to create",
                        "default": "column"
                    },
                    "data_range": {
                        "type": "string",
                        "description": "Data range for the chart (e.g., 'A1:C5')"
                    },
                    "title": {
                        "type": "string",
                        "description": "Chart title (optional)"
                    },
                    "x_axis_title": {
                        "type": "string",
                        "description": "X-axis title (optional)"
                    },
                    "y_axis_title": {
                        "type": "string",
                        "description": "Y-axis title (optional)"
                    }
                },
                "required": ["file_path"]
            }
        )
    ]


@server.call_tool()
async def handle_call_tool(name: str, arguments: dict[str, Any]) -> Sequence[TextContent | ImageContent | EmbeddedResource]:
    """Handle tool calls."""
    try:
        sheet_ops = SpreadsheetOperation()

        if name == "create_workbook":
            request = CreateWorkbookRequest(**arguments)
            result = sheet_ops.create_workbook(
                file_path=request.file_path,
                sheet_names=request.sheet_names
            )

        elif name == "write_data":
            request = WriteDataRequest(**arguments)
            result = sheet_ops.write_data(
                file_path=request.file_path,
                data=request.data,
                sheet_name=request.sheet_name,
                start_row=request.start_row,
                start_col=request.start_col,
                headers=request.headers
            )

        elif name == "read_data":
            request = ReadDataRequest(**arguments)
            result = sheet_ops.read_data(
                file_path=request.file_path,
                sheet_name=request.sheet_name,
                start_row=request.start_row,
                end_row=request.end_row,
                start_col=request.start_col,
                end_col=request.end_col
            )

        elif name == "format_cells":
            request = FormatCellsRequest(**arguments)
            result = sheet_ops.format_cells(
                file_path=request.file_path,
                cell_range=request.cell_range,
                sheet_name=request.sheet_name,
                font_name=request.font_name,
                font_size=request.font_size,
                font_bold=request.font_bold,
                font_italic=request.font_italic,
                font_color=request.font_color,
                background_color=request.background_color,
                alignment=request.alignment
            )

        elif name == "add_formula":
            request = AddFormulaRequest(**arguments)
            result = sheet_ops.add_formula(
                file_path=request.file_path,
                cell=request.cell,
                formula=request.formula,
                sheet_name=request.sheet_name
            )

        elif name == "analyze_workbook":
            request = AnalyzeWorkbookRequest(**arguments)
            result = sheet_ops.analyze_workbook(
                file_path=request.file_path,
                include_structure=request.include_structure,
                include_data_summary=request.include_data_summary,
                include_formulas=request.include_formulas
            )

        elif name == "create_chart":
            # Handle create_chart with dynamic arguments
            result = sheet_ops.create_chart(**arguments)

        else:
            result = {"success": False, "error": f"Unknown tool: {name}"}

    except Exception as e:
        logger.error(f"Error in {name}: {str(e)}")
        result = {"success": False, "error": str(e)}

    return [TextContent(type="text", text=json.dumps(result, indent=2))]


async def main():
    """Main server entry point."""
    logger.info("Starting XLSX MCP Server...")

    from mcp.server.stdio import stdio_server

    logger.info("Waiting for MCP client connection...")
    async with stdio_server() as (read_stream, write_stream):
        logger.info("MCP client connected, starting server...")
        await server.run(
            read_stream,
            write_stream,
            InitializationOptions(
                server_name="xlsx-server",
                server_version="0.1.0",
                capabilities={
                    "tools": {},
                    "logging": {},
                },
            ),
        )


if __name__ == "__main__":
    asyncio.run(main())
